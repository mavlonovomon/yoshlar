import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import Yosh


def _normalize_header(value):
    return (str(value or "").strip().lower().replace(" ", ""))


def _clean_jshshir(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            value = int(value)
    text = str(value).strip()
    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""
    if len(digits) < 14:
        digits = digits.zfill(14)
    return digits if len(digits) == 14 else ""


def _clean_guvohnoma(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if not text or text.lower() in {"none", "nan", "null", "-"}:
        return ""
    return text.upper()


class Command(BaseCommand):
    help = "sorovnoma.xlsx faylidan JSHSHIR orqali Guvohnoma raqamini Yosh jadvaliga import qiladi."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            default="sorovnoma.xlsx",
            help="Excel fayl yo'li (default: sorovnoma.xlsx)",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default="",
            help="Sheet nomi (bo'sh bo'lsa birinchi sheet olinadi)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="DBga yozmaydi, faqat natijani ko'rsatadi.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"]).resolve()
        if not file_path.exists():
            raise CommandError(f"Fayl topilmadi: {file_path}")

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sheet_name = options.get("sheet")
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(
                    f"'{sheet_name}' sheet topilmadi. Mavjudlar: {', '.join(workbook.sheetnames)}"
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_values:
            raise CommandError("Sarlavha qatori topilmadi.")

        normalized_headers = [_normalize_header(h) for h in header_values]

        try:
            jshshir_idx = normalized_headers.index("jshshir")
        except ValueError as exc:
            raise CommandError("JSHSHIR ustuni topilmadi.") from exc

        guvohnoma_idx = -1
        for idx, item in enumerate(normalized_headers):
            if item in {"guvohnomaraqami", "guvohnoma"}:
                guvohnoma_idx = idx
                break
        if guvohnoma_idx < 0:
            raise CommandError("Guvohnoma raqami ustuni topilmadi.")

        mapping = {}
        total_rows = 0
        empty_guvohnoma_rows = 0
        invalid_jshshir_rows = 0
        duplicate_conflicts = 0

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            total_rows += 1
            jshshir = _clean_jshshir(row[jshshir_idx] if jshshir_idx < len(row) else None)
            guvohnoma = _clean_guvohnoma(row[guvohnoma_idx] if guvohnoma_idx < len(row) else None)

            if not jshshir:
                invalid_jshshir_rows += 1
                continue
            if not guvohnoma:
                empty_guvohnoma_rows += 1
                continue

            prev = mapping.get(jshshir)
            if prev and prev != guvohnoma:
                duplicate_conflicts += 1
                continue
            mapping[jshshir] = guvohnoma

        if not mapping:
            self.stdout.write(self.style.WARNING("Import uchun mos satr topilmadi."))
            return

        yosh_rows = list(Yosh.objects.filter(jshshir__in=mapping.keys()).only("id", "jshshir", "guvohnoma_raqami"))
        found_in_db = len(yosh_rows)

        to_update = []
        unchanged = 0
        for item in yosh_rows:
            new_value = mapping.get(item.jshshir, "")
            if not new_value:
                continue
            if (item.guvohnoma_raqami or "") == new_value:
                unchanged += 1
                continue
            item.guvohnoma_raqami = new_value
            to_update.append(item)

        if not options.get("dry_run") and to_update:
            with transaction.atomic():
                Yosh.objects.bulk_update(to_update, ["guvohnoma_raqami"], batch_size=1000)

        not_found_in_db = len(mapping) - found_in_db
        self.stdout.write(f"Fayl: {file_path}")
        self.stdout.write(f"Sheet: {worksheet.title}")
        self.stdout.write(f"Jami satrlar: {total_rows}")
        self.stdout.write(f"Guvohnoma bor satrlar (unikal JSHSHIR): {len(mapping)}")
        self.stdout.write(f"Bo'sh guvohnoma satrlar: {empty_guvohnoma_rows}")
        self.stdout.write(f"Noto'g'ri JSHSHIR satrlar: {invalid_jshshir_rows}")
        self.stdout.write(f"Takroriy JSHSHIR ziddiyatlari: {duplicate_conflicts}")
        self.stdout.write(f"DBda topilganlar: {found_in_db}")
        self.stdout.write(f"DBda topilmaganlar: {not_found_in_db}")
        self.stdout.write(f"Yangilanadiganlar: {len(to_update)}")
        self.stdout.write(f"O'zgarmaganlar: {unchanged}")

        if options.get("dry_run"):
            self.stdout.write(self.style.WARNING("Dry-run: DBga yozilmadi."))
        else:
            self.stdout.write(self.style.SUCCESS("Guvohnoma raqamlari muvaffaqiyatli import qilindi."))
