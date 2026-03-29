import re
from collections import OrderedDict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from core.models import MaktabOquvchi, Yosh
from core.view_helpers import is_active_school_class


HEADER_ALIASES = {
    "id": "external_id",
    "полноеименование": "fullname",
    "полноенаименование": "fullname",
    "датарождения": "birth_date",
    "датарожденя": "birth_date",
    "пол": "gender",
    "jinsi": "gender",
    "millati": "nationality",
    "nationality": "nationality",
    "национальность": "nationality",
    "fuqaroligi": "citizenship",
    "гражданство": "citizenship",
    "pinfl": "pinfl",
    "пинфл": "pinfl",
    "seriya": "document_series",
    "серия": "document_series",
    "nomerdokumenta": "document_number",
    "номердокумента": "document_number",
    "tashkilot": "organization",
    "организация": "organization",
    "regionorganizacii": "organization_region",
    "регионорганизации": "organization_region",
    "klass": "klass",
    "класс": "klass",
    "class": "klass",
}

PASSPORT_SUFFIXES = {"AA", "AB", "AC", "AD", "AE", "FA", "FS"}


def _normalize_header(value):
    return re.sub(r"\s+", "", str(value or "").strip().casefold())


def _clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if text.lower() in {"none", "nan", "null", "-", "—", "–"}:
        return ""
    return text


def _clean_pinfl(value):
    text = re.sub(r"\D", "", _clean_text(value))
    return text.zfill(14) if text else ""


def _clean_int(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    return int(digits) if digits else None


def _clean_date(value):
    if not value:
        return None
    if hasattr(value, "date"):
        return value.date() if callable(value.date) else value
    return value


def _document_type_from_series(series: str) -> str:
    series = (series or "").upper().strip()
    if not series:
        return ""
    return "passport" if any(series.endswith(suffix) for suffix in PASSPORT_SUFFIXES) else "guvohnoma"


class Command(BaseCommand):
    help = "RP.xlsx faylidagi ma'lumotlarni PINFL bo'yicha Yosh bilan birlashtiradi, qolganlarini aniqlanmaganlarga yuboradi."

    def add_arguments(self, parser):
        parser.add_argument("--file", type=str, default=r"C:\Users\Genius007\Desktop\RP.xlsx", help="Excel fayl yo'li")
        parser.add_argument("--sheet", type=str, default="", help="Sheet nomi")
        parser.add_argument("--dry-run", action="store_true", help="DBga yozmaydi")
        parser.add_argument("--clear", action="store_true", help="Avval yoshlar maktab ma'lumotlarini tozalaydi")

    def handle(self, *args, **options):
        file_path = Path(options["file"]).expanduser().resolve()
        if not file_path.exists():
            raise CommandError(f"Fayl topilmadi: {file_path}")

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sheet_name = options.get("sheet")
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f"'{sheet_name}' sheet topilmadi. Mavjudlar: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_values:
            raise CommandError("Sarlavha qatori topilmadi.")

        columns = {}
        for idx, header in enumerate(header_values):
            mapped = HEADER_ALIASES.get(_normalize_header(header))
            if mapped:
                columns[mapped] = idx

        required = {"external_id", "fullname", "birth_date", "pinfl", "organization", "klass"}
        missing = sorted(required - set(columns))
        if missing:
            raise CommandError(f"Kerakli ustunlar topilmadi: {', '.join(missing)}")

        total_rows = 0
        skipped_rows = 0
        duplicate_pinfl_rows = 0
        payload_map = OrderedDict()
        now = timezone.now()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            total_rows += 1

            external_id = _clean_int(row[columns["external_id"]] if columns["external_id"] < len(row) else None)
            fullname = _clean_text(row[columns["fullname"]] if columns["fullname"] < len(row) else None)
            birth_date = _clean_date(row[columns["birth_date"]] if columns["birth_date"] < len(row) else None)
            pinfl = _clean_pinfl(row[columns["pinfl"]] if columns["pinfl"] < len(row) else None)
            organization = _clean_text(row[columns["organization"]] if columns["organization"] < len(row) else None)
            klass = _clean_text(row[columns["klass"]] if columns["klass"] < len(row) else None)

            if not external_id or not fullname or not birth_date or not pinfl or not organization:
                skipped_rows += 1
                continue

            item = {
                "external_id": external_id,
                "fullname": fullname,
                "birth_date": birth_date,
                "gender": _clean_text(row[columns["gender"]] if "gender" in columns and columns["gender"] < len(row) else None),
                "nationality": _clean_text(row[columns["nationality"]] if "nationality" in columns and columns["nationality"] < len(row) else None),
                "citizenship": _clean_text(row[columns["citizenship"]] if "citizenship" in columns and columns["citizenship"] < len(row) else None),
                "pinfl": pinfl,
                "document_series": _clean_text(row[columns["document_series"]] if "document_series" in columns and columns["document_series"] < len(row) else None),
                "document_number": _clean_text(row[columns["document_number"]] if "document_number" in columns and columns["document_number"] < len(row) else None),
                "organization": organization,
                "organization_region": _clean_text(row[columns["organization_region"]] if "organization_region" in columns and columns["organization_region"] < len(row) else None),
                "klass": klass,
            }

            if pinfl in payload_map:
                duplicate_pinfl_rows += 1
            payload_map[pinfl] = item

        payload = list(payload_map.values())
        if not payload:
            self.stdout.write(self.style.WARNING("Import uchun mos satr topilmadi."))
            return

        yosh_map = Yosh.objects.in_bulk(field_name="jshshir")
        staging_items = []

        if options.get("clear"):
            Yosh.objects.update(
                school_external_id=None,
                school_gender="",
                school_nationality="",
                school_citizenship="",
                school_document_series="",
                school_document_number="",
                school_organization="",
                school_organization_region="",
                school_class="",
                school_imported_at=None,
            )
            MaktabOquvchi.objects.all().delete()
            for yosh in yosh_map.values():
                yosh.passport_number = yosh.passport_number or ""
                yosh.guvohnoma_raqami = yosh.guvohnoma_raqami or ""

        changed_yoshes = []
        matched_rows = 0
        unmatched_rows = 0
        inactive_rows = 0
        passport_rows = 0
        guvohnoma_rows = 0
        matched_pinfls = set()
        inactive_pinfls = set()

        for item in payload:
            yosh = yosh_map.get(item["pinfl"])
            active_school = is_active_school_class(item["klass"])
            if not active_school:
                inactive_rows += 1
            if yosh:
                matched_rows += 1
                matched_pinfls.add(item["pinfl"])
                yosh.fullname = item["fullname"] or yosh.fullname
                if item["birth_date"]:
                    yosh.birth_date = item["birth_date"]
                yosh.school_external_id = item["external_id"]
                yosh.school_gender = item["gender"]
                yosh.school_nationality = item["nationality"]
                yosh.school_citizenship = item["citizenship"]
                yosh.school_document_series = item["document_series"]
                yosh.school_document_number = item["document_number"]
                yosh.school_organization = item["organization"]
                yosh.school_organization_region = item["organization_region"]
                yosh.school_class = item["klass"]
                yosh.school_imported_at = now

                doc_value = f"{item['document_series']}{item['document_number']}".strip()
                doc_type = _document_type_from_series(item["document_series"])
                if doc_value:
                    if doc_type == "passport":
                        passport_rows += 1
                        if not yosh.passport_number:
                            yosh.passport_number = doc_value
                    else:
                        guvohnoma_rows += 1
                        if not yosh.guvohnoma_raqami:
                            yosh.guvohnoma_raqami = doc_value

                changed_yoshes.append(yosh)
                continue

            if not active_school:
                inactive_pinfls.add(item["pinfl"])
                continue

            unmatched_rows += 1
            staging_items.append(
                {
                    "external_id": item["external_id"],
                    "fullname": item["fullname"],
                    "birth_date": item["birth_date"],
                    "gender": item["gender"],
                    "nationality": item["nationality"],
                    "citizenship": item["citizenship"],
                    "pinfl": item["pinfl"],
                    "document_series": item["document_series"],
                    "document_number": item["document_number"],
                    "organization": item["organization"],
                    "organization_region": item["organization_region"],
                    "klass": item["klass"],
                }
            )

        if options.get("dry_run"):
            self.stdout.write(f"Fayl: {file_path}")
            self.stdout.write(f"Sheet: {worksheet.title}")
            self.stdout.write(f"Jami satrlar: {total_rows}")
            self.stdout.write(f"Importga mos satrlar: {len(payload)}")
            self.stdout.write(f"Takroriy PINFL satrlar: {duplicate_pinfl_rows}")
            self.stdout.write(f"Yosh bilan moslashganlar: {matched_rows}")
            self.stdout.write(f"Mos kelmaganlar: {unmatched_rows}")
            self.stdout.write(f"Faol bo'lmaganlar: {inactive_rows}")
            self.stdout.write(f"Passportga tushganlar: {passport_rows}")
            self.stdout.write(f"Guvohnomaga tushganlar: {guvohnoma_rows}")
            self.stdout.write(f"Stagingga tushganlar: {len(staging_items)}")
            self.stdout.write(self.style.WARNING("Dry-run: DBga yozilmadi."))
            return

        with transaction.atomic():
            if changed_yoshes:
                Yosh.objects.bulk_update(
                    changed_yoshes,
                    [
                        "fullname",
                        "birth_date",
                        "school_external_id",
                        "school_gender",
                        "school_nationality",
                        "school_citizenship",
                        "school_document_series",
                        "school_document_number",
                        "school_organization",
                        "school_organization_region",
                        "school_class",
                        "school_imported_at",
                        "passport_number",
                        "guvohnoma_raqami",
                    ],
                    batch_size=1000,
                )

            if matched_pinfls:
                MaktabOquvchi.objects.filter(pinfl__in=matched_pinfls).delete()
            if inactive_pinfls:
                MaktabOquvchi.objects.filter(pinfl__in=inactive_pinfls).delete()

            if staging_items:
                MaktabOquvchi.objects.bulk_create(
                    [MaktabOquvchi(**item) for item in staging_items],
                    batch_size=1000,
                    update_conflicts=True,
                    unique_fields=["pinfl"],
                    update_fields=[
                        "external_id",
                        "fullname",
                        "birth_date",
                        "gender",
                        "nationality",
                        "citizenship",
                        "document_series",
                        "document_number",
                        "organization",
                        "organization_region",
                        "klass",
                        "updated_at",
                    ],
                )

        self.stdout.write(f"Fayl: {file_path}")
        self.stdout.write(f"Sheet: {worksheet.title}")
        self.stdout.write(f"Jami satrlar: {total_rows}")
        self.stdout.write(f"Importga mos satrlar: {len(payload)}")
        self.stdout.write(f"Takroriy PINFL satrlar: {duplicate_pinfl_rows}")
        self.stdout.write(f"Yosh bilan moslashganlar: {matched_rows}")
        self.stdout.write(f"Mos kelmaganlar: {unmatched_rows}")
        self.stdout.write(f"Faol bo'lmaganlar: {inactive_rows}")
        self.stdout.write(f"Passportga tushganlar: {passport_rows}")
        self.stdout.write(f"Guvohnomaga tushganlar: {guvohnoma_rows}")
        self.stdout.write(f"Stagingga tushganlar: {len(staging_items)}")
        self.stdout.write(f"Yangilandi: {len(changed_yoshes)}")
        self.stdout.write(f"O'tkazib yuborildi: {skipped_rows}")
        self.stdout.write(self.style.SUCCESS("RP ma'lumotlari Yosh jadvaliga birlashtirildi, qolganlari aniqlanmaganlarga yuborildi."))
