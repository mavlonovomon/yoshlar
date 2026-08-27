import io
import os
from datetime import datetime

from django.conf import settings
from reportlab.lib import colors as rlcolors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from .services.kpi_service import MODULE_COLUMNS, MODULE_KEYS


FONT_DIR = os.path.join(settings.STATIC_ROOT, 'fonts')
if not os.path.exists(FONT_DIR):
    FONT_DIR = os.path.join(settings.BASE_DIR, 'static', 'fonts')

try:
    pdfmetrics.registerFont(TTFont('NotoSans', os.path.join(FONT_DIR, 'NotoSans-Regular.ttf')))
    pdfmetrics.registerFont(TTFont('NotoSans-Bold', os.path.join(FONT_DIR, 'NotoSans-Bold.ttf')))
    DEFAULT_FONT = 'NotoSans'
    BOLD_FONT = 'NotoSans-Bold'
except Exception:
    DEFAULT_FONT = 'Helvetica'
    BOLD_FONT = 'Helvetica-Bold'


_TRAFFIC_FILL = {
    'green': rlcolors.HexColor('#c8e6c9'),
    'yellow': rlcolors.HexColor('#fff3cd'),
    'red': rlcolors.HexColor('#f8d7da'),
}
_EXCEL_FILL = {
    'green': PatternFill('solid', fgColor='C6EFCE'),
    'yellow': PatternFill('solid', fgColor='FFEB9C'),
    'red': PatternFill('solid', fgColor='FFC7CE'),
}


def _build_table_rows(rows, visible_keys):
    header = ['O\'rin', 'Mahalla / Yetakchi'] + [
        next(c['label'] for c in MODULE_COLUMNS if c['key'] == k) for k in visible_keys
    ] + ['Itog']
    body = [header]
    for r in rows:
        line = [
            r['rank'],
            f"{r['mahalla_name']}\n{r['leader'].full_name or r['leader'].username}",
        ]
        for k in visible_keys:
            line.append(f"{r['modules'][k]['pct']:.1f}%")
        line.append(f"{r['total_score']:.1f}")
        body.append(line)
    return body


def build_pdf(rows, visible_keys, title, subtitle):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title=title,
    )
    title_style = ParagraphStyle('t', fontName=BOLD_FONT, fontSize=15, leading=19)
    sub_style = ParagraphStyle('s', fontName=DEFAULT_FONT, fontSize=9, leading=12,
                               textColor=rlcolors.HexColor('#555555'))
    cell = ParagraphStyle('c', fontName=DEFAULT_FONT, fontSize=7, leading=9, alignment=TA_CENTER)
    cell_left = ParagraphStyle('cl', fontName=DEFAULT_FONT, fontSize=7, leading=9, alignment=TA_LEFT)
    head = ParagraphStyle('h', fontName=BOLD_FONT, fontSize=7.5, leading=10, alignment=TA_CENTER,
                          textColor=rlcolors.white)

    body = _build_table_rows(rows, visible_keys)
    ncols = len(body[0])
    data = []
    for ri, line in enumerate(body):
        if ri == 0:
            data.append([Paragraph(c, head) for c in line])
        else:
            data.append([
                Paragraph(str(c), cell_left if ci in (0, 1) else cell)
                for ci, c in enumerate(line)
            ])

    table = Table(data, repeatRows=1, colWidths=[14 * mm, 52 * mm] + [17 * mm] * (ncols - 3) + [22 * mm])
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), rlcolors.HexColor('#1f3a5f')),
        ('GRID', (0, 0), (-1, -1), 0.4, rlcolors.HexColor('#b0b8c4')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    for ri, r in enumerate(rows, start=1):
        fill = _TRAFFIC_FILL.get(r['traffic'])
        if fill:
            style.append(('BACKGROUND', (0, ri), (-1, ri), fill))
    table.setStyle(TableStyle(style))

    story = [
        Paragraph(title, title_style),
        Paragraph(subtitle, sub_style),
        Spacer(1, 6),
        table,
    ]
    doc.build(story)
    return buf.getvalue()


def build_excel(rows, visible_keys, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI"
    header = ['O\'rin', 'Mahalla', 'Yetakchi'] + [
        next(c['label'] for c in MODULE_COLUMNS if c['key'] == k) for k in visible_keys
    ] + ['Itog']
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F3A5F')
        c.alignment = Alignment(horizontal='center', vertical='center')
    for r in rows:
        ws.append([
            r['rank'],
            r['mahalla_name'],
            r['leader'].full_name or r['leader'].username,
        ] + [r['modules'][k]['pct'] for k in visible_keys] + [r['total_score']])
    fill = _EXCEL_FILL
    for ri, r in enumerate(rows, start=2):
        color = fill.get(r['traffic'])
        if color:
            for ci in range(1, len(header) + 1):
                ws.cell(row=ri, column=ci).fill = color
    thin = Side(style='thin', color='B0B8C4')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(header)):
        for c in row:
            c.border = border
    for ci in range(1, len(header) + 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
