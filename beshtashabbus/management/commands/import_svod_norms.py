"""
5tsvod.xlsx faylidagi 'svod' sheetidan norma qatorlarini import qiladi.

Foydalanish:
    python manage.py import_svod_norms 5tsvod.xlsx
"""
import openpyxl
from django.core.management.base import BaseCommand

from beshtashabbus.models import FiveInitiativeSvodNorm


class Command(BaseCommand):
    help = "5tsvod.xlsx faylidagi 'svod' sheetidan norma qatorlarini import qiladi."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Excel fayl yo'li (masalan 5tsvod.xlsx)")
        parser.add_argument(
            "--sheet",
            type=str,
            default="svod",
            help="Sheet nomi (default: svod)",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Import oldidan barcha mavjud normalarni o'chiradi",
        )

    def handle(self, *args, **options):
        file_path = options["file"]
        sheet_name = options["sheet"]

        self.stdout.write(f"Fayl yuklanmoqda: {file_path} (sheet: {sheet_name}) ...")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            self.stderr.write(f"'{sheet_name}' nomli sheet topilmadi. Mavjud sheetlar: {wb.sheetnames}")
            return

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Birinchi qator – sarlavha
        header = next(rows_iter)
        # Ustun 0: Категория выбора
        # Ustun 1: Направление выбора
        # Ustun 2: Возрастная категория
        # Ustun 3: Пол
        # Ustun 4: Norma
        # Ustun 5+: mahalla nomlari

        if options["clear"]:
            deleted, _ = FiveInitiativeSvodNorm.objects.all().delete()
            self.stdout.write(f"  {deleted} ta eski norma o'chirildi.")

        created = 0
        updated = 0
        for row_idx, row in enumerate(rows_iter, start=1):
            category = str(row[0] or "").strip()
            direction = str(row[1] or "").strip()
            age_cat = str(row[2] or "").strip()
            gender_raw = str(row[3] or "").strip() if row[3] is not None else ""
            norma = int(row[4] or 0) if row[4] is not None else 0

            if not category or not direction:
                continue

            # gender normalisation
            gender = ""
            if gender_raw:
                g = gender_raw.lower()
                if "муж" in g or "erkak" in g or "male" in g:
                    gender = "male"
                elif "жен" in g or "ayol" in g or "female" in g:
                    gender = "female"
                else:
                    gender = gender_raw

            obj, was_created = FiveInitiativeSvodNorm.objects.update_or_create(
                selection_category=category,
                direction=direction,
                age_category=age_cat,
                gender=gender,
                defaults={
                    "norma": norma,
                    "row_order": row_idx,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        self.stdout.write(
            self.style.SUCCESS(
                f"Import tugadi: {created} ta yangi, {updated} ta yangilangan norma."
            )
        )
