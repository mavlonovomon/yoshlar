from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.db.models import Count, Q
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.core.files.storage import default_storage
from datetime import datetime
import openpyxl
import logging
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

from .models import UnemployedYouth, ResponsibleLeader, YouthMeeting, AssistanceInfo
from .forms import UnemployedYouthForm, MeetingForm, AssistanceForm
from core.models import Mahalla, User, Yosh
from core.view_helpers import apply_sorting, normalize_sort_params

logger = logging.getLogger(__name__)


def _is_truthy(value):
    return str(value).strip().lower() in {'1', 'true', 'on', 'yes'}


def _wipe_assistance_data(assistance, old_document_name=None):
    """Oldingi yordam ma'lumotini to'liq tozalash (fayl bilan birga)."""
    AssistanceInfo.objects.filter(pk=assistance.pk).update(
        provided=False,
        assistance_type=None,
        date_provided=None,
        document='',
    )
    if old_document_name:
        try:
            default_storage.delete(old_document_name)
        except PermissionError:
            logger.warning("Document could not be removed due to permission lock: %s", old_document_name)


class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser or getattr(self.request.user, 'role', None) == 'RAHBAR'

class MahallaRestrictedMixin:
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if not getattr(user, 'is_site_admin', False) and user.mahalla:
            # Check if this model has a direct mahalla link or through yosh
            if hasattr(self.model, 'yosh'):
                return queryset.filter(yosh__mahalla=user.mahalla)
            elif hasattr(self.model, 'mahalla'):
                return queryset.filter(mahalla=user.mahalla)
            elif self.model == Mahalla:
                return queryset.filter(id=user.mahalla.id)
        return queryset

def _build_unemployed_totals(queryset):
    totals = queryset.aggregate(
        total_youth=Count('id'),
        with_meeting=Count('meetings', distinct=True),
        total_assisted=Count('id', filter=Q(assistance__provided=True)),
        ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
        kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
        migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
        yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
        subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
    )
    totals['not_assisted'] = totals['total_youth'] - totals['total_assisted']
    totals['percent'] = round((totals['total_assisted'] / totals['total_youth'] * 100), 1) if totals['total_youth'] > 0 else 0
    return totals


def _build_mahalla_statistics():
    stats = list(
        Mahalla.objects.annotate(
            total_youth=Count('yoshlar__unemployed_profile'),
            with_meeting=Count('yoshlar__unemployed_profile__meetings', distinct=True),
            total_assisted=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True)),
            ish=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='ISH')),
            kredit=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='KREDIT')),
            migratsiya=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='MIGRATSIYA')),
            yer=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='YER')),
            subsidiya=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='SUBSIDIYA')),
        ).order_by('name')
    )
    for row in stats:
        row.not_assisted = row.total_youth - row.total_assisted
        row.percent = round((row.total_assisted / row.total_youth * 100), 1) if row.total_youth > 0 else 0
    return stats


def _build_leader_statistics():
    stats = list(
        ResponsibleLeader.objects.annotate(
            total_youth=Count('assigned_youths'),
            with_meeting=Count('assigned_youths__meetings', distinct=True),
            total_assisted=Count('assigned_youths', filter=Q(assigned_youths__assistance__provided=True)),
            ish=Count('assigned_youths', filter=Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='ISH')),
            kredit=Count('assigned_youths', filter=Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='KREDIT')),
            migratsiya=Count('assigned_youths', filter=Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='MIGRATSIYA')),
            yer=Count('assigned_youths', filter=Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='YER')),
            subsidiya=Count('assigned_youths', filter=Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='SUBSIDIYA')),
        ).filter(total_youth__gt=0).order_by('full_name')
    )
    for row in stats:
        row.not_assisted = row.total_youth - row.total_assisted
        row.percent = round((row.total_assisted / row.total_youth * 100), 1) if row.total_youth > 0 else 0
    return stats


def _build_professional_statistics():
    queryset = UnemployedYouth.objects.select_related('leader', 'yosh__mahalla', 'assistance')
    raw_stats = queryset.values(
        'leader_id', 'leader__full_name', 'leader__position', 'leader__sector',
        'yosh__mahalla_id', 'yosh__mahalla__name'
    ).annotate(
        total_youth=Count('id'),
        with_meeting=Count('meetings', distinct=True),
        total_assisted=Count('id', filter=Q(assistance__provided=True)),
        ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
        kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
        migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
        yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
        subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
    ).order_by('yosh__mahalla__name', 'leader__full_name')

    stats = []
    for row in raw_stats:
        if row['leader_id'] is None:
            continue
        row['not_assisted'] = row['total_youth'] - row['total_assisted']
        row['percent'] = round((row['total_assisted'] / row['total_youth'] * 100), 1) if row['total_youth'] > 0 else 0
        stats.append(row)
    return stats


class SvodTabsView(LoginRequiredMixin, TemplateView):
    template_name = 'ishsiz_yoshlar/svod_tabs.html'
    default_tab = 'mahalla'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        active_tab = (self.request.GET.get('tab') or getattr(self, 'default_tab', 'mahalla') or 'mahalla').strip().lower()
        if active_tab not in {'mahalla', 'leader', 'professional'}:
            active_tab = 'mahalla'

        queryset = UnemployedYouth.objects.all()
        totals = _build_unemployed_totals(queryset)

        context.update({
            'active_tab': active_tab,
            'mahalla_statistics': _build_mahalla_statistics(),
            'leader_statistics': _build_leader_statistics(),
            'professional_statistics': _build_professional_statistics(),
            'totals': totals,
            'current_time': datetime.now(),
        })
        return context


class LeaderSvodView(LoginRequiredMixin, ListView):
    model = ResponsibleLeader
    template_name = 'ishsiz_yoshlar/leader_svod.html'
    context_object_name = 'statistics'

    def get_queryset(self):
        leaders = ResponsibleLeader.objects.all()
        filter_q = Q()

        return leaders.annotate(
            total_youth=Count('assigned_youths', filter=filter_q),
            with_meeting=Count('assigned_youths__meetings', filter=filter_q, distinct=True),
            total_assisted=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True)),
            ish=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='ISH')),
            kredit=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='KREDIT')),
            migratsiya=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='MIGRATSIYA')),
            yer=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='YER')),
            subsidiya=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='SUBSIDIYA')),
        ).filter(total_youth__gt=0).order_by('full_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = UnemployedYouth.objects.all()

        stats = []
        for s in context['statistics']:
            s.not_assisted = s.total_youth - s.total_assisted
            s.percent = round((s.total_assisted / s.total_youth * 100), 1) if s.total_youth > 0 else 0
            stats.append(s)
        context['statistics'] = stats

        totals = queryset.aggregate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        )
        totals['not_assisted'] = totals['total_youth'] - totals['total_assisted']
        totals['percent'] = round((totals['total_assisted'] / totals['total_youth'] * 100), 1) if totals['total_youth'] > 0 else 0
        context['totals'] = totals
        return context

class SvodView(LoginRequiredMixin, ListView):
    model = UnemployedYouth
    template_name = 'ishsiz_yoshlar/svod.html'
    context_object_name = 'statistics'

    def get_queryset(self):
        mahallas = Mahalla.objects.all()
        return mahallas.annotate(
            total_youth=Count('yoshlar__unemployed_profile'),
            with_meeting=Count('yoshlar__unemployed_profile__meetings', distinct=True),
            total_assisted=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True)),
            ish=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='ISH')),
            kredit=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='KREDIT')),
            migratsiya=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='MIGRATSIYA')),
            yer=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='YER')),
            subsidiya=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='SUBSIDIYA')),
        ).order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = UnemployedYouth.objects.all()

        stats = []
        for s in context['statistics']:
            s.not_assisted = s.total_youth - s.total_assisted
            s.percent = round((s.total_assisted / s.total_youth * 100), 1) if s.total_youth > 0 else 0
            stats.append(s)
        context['statistics'] = stats

        totals = queryset.aggregate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        )
        totals['not_assisted'] = totals['total_youth'] - totals['total_assisted']
        totals['percent'] = round((totals['total_assisted'] / totals['total_youth'] * 100), 1) if totals['total_youth'] > 0 else 0
        context['totals'] = totals
        return context

class DetailedSvodView(LoginRequiredMixin, TemplateView):
    template_name = 'ishsiz_yoshlar/detailed_svod.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Base queryset
        queryset = UnemployedYouth.objects.select_related('leader', 'yosh__mahalla', 'assistance')

        # Process statistics grouped by Leader and Mahalla
        # In this specific app logic, we group by (leader, mahalla)
        raw_stats = queryset.values(
            'leader_id', 'leader__full_name', 'leader__position', 'leader__level', 'leader__sector',
            'yosh__mahalla_id', 'yosh__mahalla__name'
        ).annotate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        ).order_by('yosh__mahalla__name', 'leader__full_name')

        processed_stats = []
        for s in raw_stats:
            if s['leader_id'] is None:
                continue
                
            total = s['total_youth']
            assisted = s['total_assisted']
            not_assisted = total - assisted
            percent = round((assisted / total * 100), 1) if total > 0 else 0
            
            s['not_assisted'] = not_assisted
            s['percent'] = percent
            processed_stats.append(s)

        # Calculate Totals for the header
        totals = queryset.aggregate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        )
        totals['not_assisted'] = totals['total_youth'] - totals['total_assisted']
        totals['percent'] = round((totals['total_assisted'] / totals['total_youth'] * 100), 1) if totals['total_youth'] > 0 else 0

        context['statistics'] = processed_stats
        context['totals'] = totals
        return context


class YoshAutocompleteView(LoginRequiredMixin, View):
    def get(self, request):
        q = (request.GET.get('q') or '').strip()
        if len(q) < 2:
            return JsonResponse({'results': []})

        queryset = Yosh.objects.select_related('mahalla').filter(unemployed_profile__isnull=True)
        user = request.user
        if not getattr(user, 'is_site_admin', False) and user.mahalla_id:
            queryset = queryset.filter(mahalla_id=user.mahalla_id)

        queryset = queryset.filter(
            Q(fullname__icontains=q) |
            Q(passport_number__icontains=q) |
            Q(jshshir__icontains=q) |
            Q(phone_number__icontains=q)
        ).only(
            'id', 'fullname', 'passport_number', 'jshshir', 'mahalla__name'
        ).order_by('fullname')[:20]

        results = []
        for yosh in queryset:
            passport = yosh.passport_number or '-'
            results.append({
                'id': yosh.id,
                'text': f"{yosh.fullname} | {passport} | {yosh.jshshir} | {yosh.mahalla.name}",
            })
        return JsonResponse({'results': results})


class UnemployedYouthListView(LoginRequiredMixin, ListView):
    model = UnemployedYouth
    template_name = 'ishsiz_yoshlar/list.html'
    context_object_name = 'youths'
    paginate_by = 25

    def get_queryset(self):
        user = self.request.user
        queryset = UnemployedYouth.objects.select_related(
            'yosh__mahalla', 'leader', 'assistance'
        ).annotate(
            meeting_count=Count('meetings', distinct=True)
        ).only(
            'id',
            'category',
            'yosh__id',
            'yosh__fullname',
            'yosh__photo',
            'yosh__phone_number',
            'yosh__passport_number',
            'yosh__jshshir',
            'yosh__mahalla__id',
            'yosh__mahalla__name',
            'leader__id',
            'leader__full_name',
            'leader__position',
            'leader__phone_number',
            'leader__organization',
            'leader__sector',
            'leader__level',
            'assistance__id',
            'assistance__provided',
        )
        
        if not getattr(user, 'is_site_admin', False) and user.mahalla:
            queryset = queryset.filter(yosh__mahalla=user.mahalla)
        
        search = self.request.GET.get('q')
        if search:
            search = search.strip()
            queryset = queryset.filter(
                Q(yosh__fullname__icontains=search) |
                Q(yosh__passport_number__icontains=search) |
                Q(yosh__jshshir__startswith=search) |
                Q(yosh__phone_number__icontains=search)
            )

        mahalla_id = self.request.GET.get('mahalla')
        if mahalla_id:
            queryset = queryset.filter(yosh__mahalla_id=mahalla_id)

        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category=category)

        leader_id = self.request.GET.get('leader')
        if leader_id:
            queryset = queryset.filter(leader_id=leader_id)

        sort_field, sort_direction = normalize_sort_params(
            self.request,
            {
                'fullname',
                'passport_number',
                'mahalla',
                'category',
                'leader',
                'meeting_count',
            },
            'fullname',
        )
        sort_map = {
            'fullname': 'yosh__fullname',
            'passport_number': 'yosh__passport_number',
            'mahalla': 'yosh__mahalla__name',
            'category': 'category',
            'leader': 'leader__full_name',
            'meeting_count': 'meeting_count',
        }
        self.sort_field = sort_field
        self.sort_direction = sort_direction
        return apply_sorting(queryset, sort_field, sort_direction, sort_map, 'fullname')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Safely cast IDs to integers for template comparison
        try:
            mahalla_id = self.request.GET.get('mahalla')
            context['selected_mahalla'] = int(mahalla_id) if mahalla_id else None
        except (ValueError, TypeError):
            context['selected_mahalla'] = None

        try:
            leader_id = self.request.GET.get('leader')
            context['selected_leader'] = int(leader_id) if leader_id else None
        except (ValueError, TypeError):
            context['selected_leader'] = None

        context['selected_category'] = self.request.GET.get('category')
        
        if getattr(user, 'is_site_admin', False):
            context['mahallas'] = Mahalla.objects.all()
        else:
            context['mahallas'] = Mahalla.objects.filter(id=user.mahalla.id) if user.mahalla else Mahalla.objects.none()
            
        context['leaders'] = ResponsibleLeader.objects.all()
        context['categories'] = UnemployedYouth.CATEGORY_CHOICES
        context['sort_field'] = getattr(self, 'sort_field', 'fullname')
        context['sort_direction'] = getattr(self, 'sort_direction', 'asc')
        return context

class UnemployedYouthCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = UnemployedYouth
    form_class = UnemployedYouthForm
    template_name = 'ishsiz_yoshlar/form.html'
    success_url = reverse_lazy('ishsiz_yoshlar:list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        user = self.request.user
        # For non-superadmins, force the youth's mahalla to match user's mahalla?
        # Actually the UnemployedYouth points to Yosh, and Yosh has mahalla.
        # Let's ensure the chosen Yosh belongs to the user's mahalla.
        yosh = form.cleaned_data.get('yosh')
        if not getattr(user, 'is_site_admin', False) and user.mahalla and yosh.mahalla != user.mahalla:
            messages.error(self.request, "Siz faqat o'zingizning mahallangizdagi yoshlarni qo'shishingiz mumkin.")
            return self.form_invalid(form)
        return super().form_valid(form)

class UnemployedYouthUpdateView(LoginRequiredMixin, AdminRequiredMixin, MahallaRestrictedMixin, UpdateView):
    model = UnemployedYouth
    form_class = UnemployedYouthForm
    template_name = 'ishsiz_yoshlar/form.html'
    success_url = reverse_lazy('ishsiz_yoshlar:list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

class UnemployedYouthDeleteView(LoginRequiredMixin, AdminRequiredMixin, MahallaRestrictedMixin, DeleteView):
    model = UnemployedYouth
    template_name = 'ishsiz_yoshlar/confirm_delete.html'
    success_url = reverse_lazy('ishsiz_yoshlar:list')

class UnemployedYouthDetailView(LoginRequiredMixin, MahallaRestrictedMixin, DetailView):
    model = UnemployedYouth
    template_name = 'ishsiz_yoshlar/detail.html'
    context_object_name = 'youth'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['meetings'] = self.object.meetings.all().order_by('-meeting_date')
        context['assistance'] = getattr(self.object, 'assistance', None)
        return context

class MeetingCreateView(LoginRequiredMixin, MahallaRestrictedMixin, CreateView):
    model = YouthMeeting
    form_class = MeetingForm
    template_name = 'ishsiz_yoshlar/meeting_form.html'

    def dispatch(self, request, *args, **kwargs):
        youth = get_object_or_404(UnemployedYouth, pk=self.kwargs['pk'])
        user = request.user
        if not getattr(user, 'is_site_admin', False) and user.mahalla and youth.yosh.mahalla != user.mahalla:
            messages.error(request, "Sizda bu yoshga kirish huquqi yo'q.")
            return redirect('ishsiz_yoshlar:list')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['youth'] = get_object_or_404(UnemployedYouth, pk=self.kwargs['pk'])
        return context

    def form_valid(self, form):
        youth = get_object_or_404(UnemployedYouth, pk=self.kwargs['pk'])
        form.instance.unemployed_youth = youth
        messages.success(self.request, "Uchrashuv muvaffaqiyatli qo'shildi.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('ishsiz_yoshlar:detail', kwargs={'pk': self.kwargs['pk']})

class AssistanceUpdateView(LoginRequiredMixin, MahallaRestrictedMixin, View):
    def dispatch(self, request, *args, **kwargs):
        youth = get_object_or_404(UnemployedYouth, pk=kwargs['pk'])
        user = request.user
        if not getattr(user, 'is_site_admin', False) and user.mahalla and youth.yosh.mahalla != user.mahalla:
            messages.error(request, "Sizda bu yoshga kirish huquqi yo'q.")
            return redirect('ishsiz_yoshlar:list')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        youth = get_object_or_404(UnemployedYouth, pk=pk)
        assistance, _ = AssistanceInfo.objects.get_or_create(unemployed_youth=youth)
        form = AssistanceForm(instance=assistance)
        return render(request, 'ishsiz_yoshlar/assistance_form.html', {
            'form': form,
            'youth': youth,
            'has_existing_assistance': bool(assistance.provided),
        })

    def post(self, request, pk):
        youth = get_object_or_404(UnemployedYouth, pk=pk)
        assistance, _ = AssistanceInfo.objects.get_or_create(unemployed_youth=youth)
        has_existing_assistance = bool(assistance.provided)
        old_document_name = assistance.document.name if has_existing_assistance and assistance.document else None
        requested_provided = _is_truthy(request.POST.get('provided'))
        replace_confirmed = _is_truthy(request.POST.get('replace_confirmed'))

        form = AssistanceForm(request.POST, request.FILES, instance=assistance)

        if has_existing_assistance and requested_provided and not replace_confirmed:
            warning_text = (
                "Bu yosh uchun avvalgi yordam ma'lumoti mavjud. "
                "Yangi yordamni saqlash uchun avvalgisini o'chirishni tasdiqlang."
            )
            form.add_error(None, warning_text)
            messages.warning(request, warning_text)
            return render(request, 'ishsiz_yoshlar/assistance_form.html', {
                'form': form,
                'youth': youth,
                'has_existing_assistance': has_existing_assistance,
            })

        if form.is_valid():
            replaced = has_existing_assistance and form.cleaned_data.get('provided')
            if replaced:
                _wipe_assistance_data(assistance, old_document_name=old_document_name)
            form.save()
            if replaced:
                messages.success(request, "Avvalgi yordam ma'lumoti o'chirildi va yangisi saqlandi.")
            else:
                messages.success(request, "Yordam ma'lumoti muvaffaqiyatli saqlandi.")
            return redirect('ishsiz_yoshlar:detail', pk=pk)
        return render(request, 'ishsiz_yoshlar/assistance_form.html', {
            'form': form,
            'youth': youth,
            'has_existing_assistance': has_existing_assistance,
        })

class MeetingUpdateView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request, pk):
        user = request.user
        queryset = UnemployedYouth.objects.all()
        if not getattr(user, 'is_site_admin', False) and user.mahalla:
            queryset = queryset.filter(yosh__mahalla=user.mahalla)
            
        youth = get_object_or_404(queryset, pk=pk)
        meeting_form = MeetingForm()
        assistance_form = AssistanceForm(instance=getattr(youth, 'assistance', None))
        return render(request, 'ishsiz_yoshlar/meeting_update.html', {
            'youth': youth,
            'meeting_form': meeting_form,
            'assistance_form': assistance_form
        })

    def post(self, request, pk):
        user = request.user
        queryset = UnemployedYouth.objects.all()
        if not getattr(user, 'is_site_admin', False) and user.mahalla:
            queryset = queryset.filter(yosh__mahalla=user.mahalla)
            
        youth = get_object_or_404(queryset, pk=pk)
        # Handle which form was submitted or both
        if 'meeting_submit' in request.POST:
            meeting_form = MeetingForm(request.POST, request.FILES)
            if meeting_form.is_valid():
                meeting = meeting_form.save(commit=False)
                meeting.unemployed_youth = youth
                meeting.save()
                messages.success(request, "Uchrashuv saqlandi.")
        
        if 'assistance_submit' in request.POST:
            assistance, created = AssistanceInfo.objects.get_or_create(unemployed_youth=youth)
            has_existing_assistance = bool(assistance.provided)
            old_document_name = assistance.document.name if has_existing_assistance and assistance.document else None
            requested_provided = _is_truthy(request.POST.get('provided'))
            replace_confirmed = _is_truthy(request.POST.get('replace_confirmed'))

            if has_existing_assistance and requested_provided and not replace_confirmed:
                messages.warning(
                    request,
                    "Avvalgi yordam ma'lumoti mavjud. Almashtirish uchun tasdiqni "
                    "\"Yordam ma'lumoti\" oynasida bering."
                )
                return redirect('ishsiz_yoshlar:assistance_update', pk=pk)

            assistance_form = AssistanceForm(request.POST, request.FILES, instance=assistance)
            if assistance_form.is_valid():
                replaced = has_existing_assistance and assistance_form.cleaned_data.get('provided')
                if replaced:
                    _wipe_assistance_data(assistance, old_document_name=old_document_name)
                assistance_form.save()
                if replaced:
                    messages.success(request, "Avvalgi yordam o'chirildi va yangisi saqlandi.")
                else:
                    messages.success(request, "Yordam ma'lumoti yangilandi.")
            else:
                messages.error(request, "Yordam ma'lumoti saqlanmadi. Maydonlarni tekshiring.")

        return redirect('ishsiz_yoshlar:list')

class ProfessionalSvodView(LoginRequiredMixin, TemplateView):
    template_name = 'ishsiz_yoshlar/professional_svod.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        queryset = UnemployedYouth.objects.select_related('leader', 'yosh__mahalla', 'assistance')

        # Statistical grouping by Leader and Mahalla
        raw_stats = queryset.values(
            'leader_id', 'leader__full_name', 'leader__position', 'leader__sector',
            'yosh__mahalla_id', 'yosh__mahalla__name'
        ).annotate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        ).order_by('yosh__mahalla__name', 'leader__full_name')

        stats = []
        for s in raw_stats:
            if s['leader_id'] is None: continue
            
            s['not_assisted'] = s['total_youth'] - s['total_assisted']
            s['percent'] = round((s['total_assisted'] / s['total_youth'] * 100), 1) if s['total_youth'] > 0 else 0
            stats.append(s)

        totals = queryset.aggregate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        )
        totals['not_assisted'] = totals['total_youth'] - totals['total_assisted']
        totals['percent'] = round((totals['total_assisted'] / totals['total_youth'] * 100), 1) if totals['total_youth'] > 0 else 0

        context['statistics'] = stats
        context['totals'] = totals
        context['current_time'] = datetime.now()
        return context

class ExportProfessionalSvodView(LoginRequiredMixin, View):
    def get(self, request):
        queryset = UnemployedYouth.objects.select_related('leader', 'yosh__mahalla', 'assistance')

        raw_stats = queryset.values(
            'leader_id', 'leader__full_name', 'leader__position', 'leader__sector',
            'yosh__mahalla_id', 'yosh__mahalla__name'
        ).annotate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        ).order_by('yosh__mahalla__name', 'leader__full_name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Svod"

        # Styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, size=11)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        # Header Title
        ws.merge_cells('A1:O1')
        ws['A1'] = "Хоразм вилояти Ҳазорасп туманида Ишсиз ёшлар бандлиги таъминлаш юзасидан амалга оширилган ишлар тўғрисида МАЪЛУМОТ"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40

        # Subheaders
        headers = [
            "№", "Туман (шаҳар) номи", "Бириктирилган маҳалла", "Масъул раҳбар (Ф.И.Ш, Лавозими)", "Соҳалар",
            "Ишсиз бириктирилган ёшлар сони", "Учрашув ўтказилган ёшлар сони", 
            "Жами бандлиги таъминланган ёшлар сони", "Доимий ишга жойлашган", "Тадбиркорлик ва кредит", 
            "Тартибли миграция", "Экин ер майдони", "Асбоб-ускуна ажратилган", "Бандлиги таъминланмаган", "Ижро ҳолати (%)"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill

        ws.row_dimensions[3].height = 60

        # Totals Calculation
        totals = queryset.aggregate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        )
        total_assisted_val = totals['total_assisted']
        total_youth_val = totals['total_youth']
        total_not_assisted = total_youth_val - total_assisted_val
        total_percent = round((total_assisted_val / total_youth_val * 100), 1) if total_youth_val > 0 else 0

        # Totals Row
        total_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        total_data = [
            "", "Ҳазорасп тумани жами", "", "", "",
            total_youth_val, totals['with_meeting'], total_assisted_val,
            totals['ish'], totals['kredit'], totals['migratsiya'], totals['yer'], totals['subsidiya'],
            total_not_assisted, f"{total_percent}%"
        ]
        ws.merge_cells('B4:E4')
        for col, val in enumerate(total_data, 1):
            cell = ws.cell(row=4, column=col)
            # Safe write: only write if it's not a MergedCell (which means it's the top-left or a normal cell)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = val
            
            # Styles must be applied to all cells in the range for borders to work correctly
            cell.font = header_font
            cell.border = thin_border
            cell.fill = total_fill
            cell.alignment = center_alignment

        # Data
        row_num = 5
        for idx, s in enumerate(raw_stats, 1):
            if s['leader_id'] is None: continue
            
            total = s['total_youth']
            assisted = s['total_assisted']
            not_assisted = total - assisted
            percent = round((assisted / total * 100), 1) if total > 0 else 0

            data = [
                idx, "Ҳазорасп тумани", s['yosh__mahalla__name'], 
                f"{s['leader__full_name']}\n{s['leader__position']}", s['leader__sector'],
                total, s['with_meeting'], assisted,
                s['ish'], s['kredit'], s['migratsiya'], s['yer'], s['subsidiya'],
                not_assisted, f"{percent}%"
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = val
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center' if col != 4 else 'left')
            
            row_num += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        for col in 'FGHIJKLMNO':
            ws.column_dimensions[col].width = 12

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=ishsiz_yoshlar_svod_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        return response

class ExportMahallaSvodView(LoginRequiredMixin, View):
    def get(self, request):
        queryset = UnemployedYouth.objects.all()
        mahallas = Mahalla.objects.all()

        stats = mahallas.annotate(
            total_youth=Count('yoshlar__unemployed_profile'),
            with_meeting=Count('yoshlar__unemployed_profile__meetings', distinct=True),
            total_assisted=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True)),
            ish=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='ISH')),
            kredit=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='KREDIT')),
            migratsiya=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='MIGRATSIYA')),
            yer=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='YER')),
            subsidiya=Count('yoshlar__unemployed_profile', filter=Q(yoshlar__unemployed_profile__assistance__provided=True, yoshlar__unemployed_profile__assistance__assistance_type='SUBSIDIYA')),
        ).order_by('name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mahalla Svod"

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        ws.merge_cells('A1:L1')
        ws['A1'] = "Ҳазорасп туманидаги маҳаллалар кесимида ишисиз ёшлар бандлиги таҳлили"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_alignment

        headers = ["№", "Маҳалла номи", "Ишсиз ёшлар сони", "Учрашув ўтказилганлар", "Жами бандлиги таъминланганлар", "Доимий ишга жойлашган", "Тадбиркорлик учун кредит ажратилган", "Тартибли миграцияга юборилган", "Экин ер майдони ажратилган", "Асбоб-ускуна ажратилган", "Бандлиги таъминланмаганлар", "Ижро ҳолати (%)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = h
            cell.font = header_font
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = center_alignment

        totals = queryset.aggregate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        )
        total_youth_val = totals['total_youth']
        total_percent = round((totals['total_assisted'] / total_youth_val * 100), 1) if total_youth_val > 0 else 0
        
        totals_row = ["", "Ҳазорасп тумани жами", total_youth_val, totals['with_meeting'], totals['total_assisted'], totals['ish'], totals['kredit'], totals['migratsiya'], totals['yer'], totals['subsidiya'], total_youth_val-totals['total_assisted'], f"{total_percent}%"]
        for col, val in enumerate(totals_row, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = val
            cell.font = header_font
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = center_alignment

        for idx, s in enumerate(stats, 1):
            assisted = s.total_assisted
            percent = round((assisted / s.total_youth * 100), 1) if s.total_youth > 0 else 0
            row = [idx, s.name, s.total_youth, s.with_meeting, assisted, s.ish, s.kredit, s.migratsiya, s.yer, s.subsidiya, s.total_youth-assisted, f"{percent}%"]
            for col, v in enumerate(row, 1):
                cell = ws.cell(row=idx+4, column=col)
                cell.value = v
                cell.border = thin_border
                cell.alignment = center_alignment

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=mahalla_svod_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        return response

class ExportLeaderSvodView(LoginRequiredMixin, View):
    def get(self, request):
        leaders = ResponsibleLeader.objects.all()
        queryset = UnemployedYouth.objects.all()
        
        filter_q = Q()

        stats = leaders.annotate(
            total_youth=Count('assigned_youths', filter=filter_q),
            with_meeting=Count('assigned_youths__meetings', filter=filter_q, distinct=True),
            total_assisted=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True)),
            ish=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='ISH')),
            kredit=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='KREDIT')),
            migratsiya=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='MIGRATSIYA')),
            yer=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='YER')),
            subsidiya=Count('assigned_youths', filter=filter_q & Q(assigned_youths__assistance__provided=True, assigned_youths__assistance__assistance_type='SUBSIDIYA')),
        ).filter(total_youth__gt=0).order_by('full_name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leader Svod"

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        ws.merge_cells('A1:M1')
        ws['A1'] = "Ҳазорасп туманидаги масъул раҳбарлар кесимида ишисиз ёшлар бандлиги таҳлили"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_alignment

        headers = ["№", "Масъул раҳбар", "Соҳаси", "Ишсиз ёшлар сони", "Учрашув", "Жами бандлиги таъминланганлар", "Доимий ишга жойлашган", "Тадбиркорлик учун кредит ажратилган", "Тартибли миграцияга юборилган", "Экин ер майдони ажратилган", "Асбоб-ускуна ажратилган", "Бандлиги таъминланмаганлар", "Ижро ҳолати (%)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = h
            cell.font = header_font
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = center_alignment

        totals = queryset.aggregate(
            total_youth=Count('id'),
            with_meeting=Count('meetings', distinct=True),
            total_assisted=Count('id', filter=Q(assistance__provided=True)),
            ish=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='ISH')),
            kredit=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='KREDIT')),
            migratsiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='MIGRATSIYA')),
            yer=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='YER')),
            subsidiya=Count('id', filter=Q(assistance__provided=True, assistance__assistance_type='SUBSIDIYA')),
        )
        total_youth_val = totals['total_youth']
        total_percent = round((totals['total_assisted'] / total_youth_val * 100), 1) if total_youth_val > 0 else 0

        totals_row = ["", "Ҳазорасп тумани жами", "", totals['total_youth'], totals['with_meeting'], totals['total_assisted'], totals['ish'], totals['kredit'], totals['migratsiya'], totals['yer'], totals['subsidiya'], total_youth_val-totals['total_assisted'], f"{total_percent}%"]
        for col, val in enumerate(totals_row, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = val
            cell.font = header_font
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = center_alignment

        for idx, s in enumerate(stats, 1):
            assisted = s.total_assisted
            percent = round((assisted / s.total_youth * 100), 1) if s.total_youth > 0 else 0
            row = [idx, f"{s.full_name}\n({s.position or ''})", s.sector or '', s.total_youth, s.with_meeting, assisted, s.ish, s.kredit, s.migratsiya, s.yer, s.subsidiya, s.total_youth-assisted, f"{percent}%"]
            for col, v in enumerate(row, 1):
                cell = ws.cell(row=idx+4, column=col)
                cell.value = v
                cell.border = thin_border
                cell.alignment = center_alignment

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=leader_svod_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        wb.save(response)
        return response

from .views_tasks import (
    NotificationListView,
    MarkNotificationReadView,
    TaskAcceptView,
    TaskCreateView,
    TaskDeleteView,
    TaskDetailView,
    TaskListView,
    TaskResponseCreateView,
    TaskReviewView,
    TaskUpdateView,
)
