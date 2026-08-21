import openpyxl
from django.core.management.base import BaseCommand
from core.models import Yosh
from ishsiz_yoshlar.models import UnemployedYouth


class Command(BaseCommand):
    help = "Excel fayldagi E (ta'lim tashkiloti) va F (yo'nalish) ustunlarini import qilish"

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Excel fayl yoli')

    def handle(self, *args, **options):
        excel_path = options['excel_path']

        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Номма-ном']

        updated = 0
        not_found_yosh = 0
        not_found_uy = 0
        skipped = 0

        for row_idx in range(4, ws.max_row + 1):
            jshshir = str(ws.cell(row=row_idx, column=4).value or '').strip()
            if not jshshir or len(jshshir) < 10:
                skipped += 1
                continue

            talim_tashkiloti = str(ws.cell(row=row_idx, column=5).value or '').strip()
            yo_nalishi = str(ws.cell(row=row_idx, column=6).value or '').strip()

            try:
                yosh = Yosh.objects.get(jshshir=jshshir)
            except Yosh.DoesNotExist:
                not_found_yosh += 1
                continue

            uy = UnemployedYouth.objects.filter(yosh=yosh).first()
            if not uy:
                not_found_uy += 1
                continue

            uy.otm_name = talim_tashkiloti
            uy.direction = yo_nalishi
            uy.save(update_fields=['otm_name', 'direction'])
            updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"\n=== IMPORT NATIJALARI ===\n"
            f"Yangilandi: {updated}\n"
            f"Yosh topilmadi: {not_found_yosh}\n"
            f"UnemployedYouth topilmadi: {not_found_uy}\n"
            f"O'tkazib yuborildi: {skipped}"
        ))
