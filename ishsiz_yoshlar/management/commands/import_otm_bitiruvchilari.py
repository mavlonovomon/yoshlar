import openpyxl
from openpyxl.styles import PatternFill
from django.core.management.base import BaseCommand
from core.models import Yosh, Mahalla
from ishsiz_yoshlar.models import UnemployedYouth, ResponsibleLeader


class Command(BaseCommand):
    help = 'OTM bitiruvchilarini Excel fayldan import qilish'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Excel fayl yoli')
        parser.add_argument('--output', type=str, default='', help='Natija Excel fayl yoli')

    def handle(self, *args, **options):
        excel_path = options['excel_path']
        output_path = options['output'] or excel_path.replace('.xlsx', '_natija.xlsx')

        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Номма-ном']

        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        created = 0
        updated = 0
        not_found = 0
        mahalla_mismatch = 0

        for row_idx in range(4, ws.max_row + 1):
            jshshir = str(ws.cell(row=row_idx, column=4).value or '').strip()
            if not jshshir or len(jshshir) < 10:
                continue

            excel_mahalla_name = str(ws.cell(row=row_idx, column=2).value or '').strip()
            fio = str(ws.cell(row=row_idx, column=3).value or '').strip()
            otm = str(ws.cell(row=row_idx, column=5).value or '').strip()
            direction = str(ws.cell(row=row_idx, column=6).value or '').strip()

            try:
                yosh = Yosh.objects.get(jshshir=jshshir)
            except Yosh.DoesNotExist:
                not_found += 1
                self.stdout.write(self.style.WARNING(
                    f"JSHSHIR topilmadi: {jshshir} ({fio})"
                ))
                continue

            excel_mahalla_norm = excel_mahalla_name.lower().replace('-', ' ').strip()
            db_mahalla_name = yosh.mahalla.name.lower().replace('-', ' ').strip() if yosh.mahalla else ''
            mahalla_match = excel_mahalla_norm == db_mahalla_name

            if not mahalla_match:
                mahalla_mismatch += 1
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col).fill = red_fill

            leader = None
            leader_name = str(ws.cell(row=row_idx, column=11).value or '').strip()
            if leader_name:
                last_part = leader_name.split('\n')[-1].strip()
                leader = ResponsibleLeader.objects.filter(
                    full_name__icontains=last_part
                ).first()

            obj, was_created = UnemployedYouth.objects.update_or_create(
                yosh=yosh,
                defaults={
                    'year': 2026,
                    'category': 'OLIY',
                    'leader': leader,
                    'otm_name': otm,
                    'direction': direction,
                }
            )
            if was_created:
                created += 1
            else:
                updated += 1

            if mahalla_match:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col).fill = green_fill

        wb.save(output_path)

        self.stdout.write(self.style.SUCCESS(
            f"\n=== IMPORT NATIJALARI ===\n"
            f"Yangi qo'shildi: {created}\n"
            f"Yangilandi (mavjud): {updated}\n"
            f"JSHSHIR topilmadi: {not_found}\n"
            f"Mahalla mos kelmasligi: {mahalla_mismatch}\n"
            f"Natija fayl: {output_path}"
        ))
